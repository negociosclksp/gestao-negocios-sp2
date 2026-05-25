# -*- coding: utf-8 -*-
import sys, re, json, os
from openpyxl import load_workbook
from datetime import datetime

# ── CAMINHOS ─────────────────────────────────────────────────────────────────
PASTA        = r"C:\Users\e5003419\Documents\Team-Dashboard"
HTML         = os.path.join(PASTA, "index.html")
EXCEL_FERIAS = os.path.join(PASTA, "F\u00c9RIAS_2025-2026.xlsx")
EXCEL_ANIV   = os.path.join(PASTA, "FolgaAniverssario.xlsx")
# Busca PDI pelo listdir (evita problema de encoding no nome)
EXCEL_PDI = None
for _f in os.listdir(PASTA):
    if _f.lower().endswith('.xlsx') and 'pid' in _f.lower() and '2026' in _f:
        EXCEL_PDI = os.path.join(PASTA, _f)
        break
if not EXCEL_PDI:
    print("PDI nao encontrado. Arquivos xlsx na pasta:")
    for _f in os.listdir(PASTA):
        if _f.lower().endswith('.xlsx'):
            print(f"  -> {_f}")
    EXCEL_PDI = os.path.join(PASTA, "PDI_NAO_ENCONTRADO.xlsx")
EXCEL_ESCALA = os.path.join(PASTA, "Escala_2026.xlsx")

# Verifica arquivos
erros = []
for nome, path in [("index.html", HTML), ("Ferias", EXCEL_FERIAS),
                   ("Aniversarios", EXCEL_ANIV), ("PDI", EXCEL_PDI), ("Escala", EXCEL_ESCALA)]:
    if os.path.exists(path):
        print(f"[OK] {nome}: {path}")
    else:
        print(f"[FALTA] {nome}: {path}")
        erros.append(nome)

if erros:
    print(f"\nERRO: Arquivos nao encontrados: {erros}")
    print("Verifique os nomes dos arquivos na pasta e ajuste o script.")
    sys.exit(1)

print()
with open(HTML, "r", encoding="utf-8") as f:
    content = f.read()

ts = datetime.now().strftime("%d/%m/%Y %H:%M")

# ── FERIAS ────────────────────────────────────────────────────────────────────
def parse_range(s):
    if not s: return "", ""
    s = str(s).strip()
    for sep in [" a ", " \u00e0 ", " - "]:
        if sep in s:
            parts = s.split(sep)
            try:
                d1 = datetime.strptime(parts[0].strip(), "%d/%m/%Y")
                d2 = datetime.strptime(parts[1].strip(), "%d/%m/%Y")
                return d1.strftime("%Y-%m-%d"), d2.strftime("%Y-%m-%d")
            except: pass
    return "", ""

wb = load_workbook(EXCEL_FERIAS, read_only=True)
ws = wb.active
ferias_lines = []
for row in list(ws.iter_rows(values_only=True))[1:]:
    drt,nome,cargo,f1,f2,obs,pend,agend = row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7]
    if not nome: continue
    s1,e1 = parse_range(f1)
    s2,e2 = parse_range(f2)
    dias  = int(pend) if pend else 0
    ag    = str(agend).strip() if agend else ""
    obs_s = str(obs).replace("'", " ").strip() if obs else ""
    nome_s = str(nome).replace("'", " ").strip()
    ferias_lines.append("{DRT:'" + str(drt).strip() + "',NOME:'" + nome_s + "',CARGO:'" + str(cargo or "").strip() + "',FERIAS1_INICIO:'" + s1 + "',FERIAS1_FIM:'" + e1 + "',FERIAS2_INICIO:'" + s2 + "',FERIAS2_FIM:'" + e2 + "',OBS:'" + obs_s + "',dias_pendentes:" + str(dias) + ",AGENDADA:'" + ag + "'},")

content = re.sub(r"// @@DATA_START@@.*?// @@DATA_END@@", "// @@DATA_START@@\n" + "\n".join(ferias_lines) + "\n// @@DATA_END@@", content, flags=re.DOTALL)
content = re.sub(r"// LAST_UPDATE: .*", "// LAST_UPDATE: " + ts, content)
content = re.sub(r"const ts = \'.*?\';", "const ts = \'" + ts + "\';", content)
print(f"Ferias: {len(ferias_lines)} colaboradores")

# ── ANIVERSARIOS ──────────────────────────────────────────────────────────────
wb2 = load_workbook(EXCEL_ANIV, read_only=True)
ws2 = wb2.active
aniv_lines = []
for row in list(ws2.iter_rows(values_only=True))[1:]:
    nome,aniv,mes,status,data_folga,coment = row[0],row[1],row[2],row[3],row[4],row[5]
    if not nome: continue
    aniv_s  = aniv.strftime("%Y-%m-%d") if hasattr(aniv,"strftime") else ""
    folga_s = data_folga.strftime("%Y-%m-%d") if hasattr(data_folga,"strftime") else ""
    nome_s   = str(nome).replace("'", " ").strip()
    mes_s    = str(mes or "").replace("'", " ").strip()
    status_s = str(status or "").replace("'", " ").strip()
    coment_s = str(coment or "").replace("'", " ").replace('"', "").strip() if coment else ""
    aniv_lines.append("{NOME:'" + nome_s + "',ANIVERSARIO:'" + aniv_s + "',MES:'" + mes_s + "',STATUS:'" + status_s + "',DATA_FOLGA:'" + folga_s + "',COMENTARIOS:'" + coment_s + "'},")

content = re.sub(r"// @@ANIV_DATA_START@@.*?// @@ANIV_DATA_END@@", "// @@ANIV_DATA_START@@\n" + "\n".join(aniv_lines) + "\n// @@ANIV_DATA_END@@", content, flags=re.DOTALL)
content = re.sub(r"// ANIV_LAST_UPDATE: .*", "// ANIV_LAST_UPDATE: " + ts, content)
print(f"Aniversarios: {len(aniv_lines)} colaboradores")

# ── PDI ───────────────────────────────────────────────────────────────────────
def esc(s):
    return str(s).replace("'", "\\'").replace("`", "\\`").replace("\n", " ").strip()

wb3 = load_workbook(EXCEL_PDI, read_only=True, data_only=True)
sheets_pdi = ["BRUNO_SARAIVA","CAMILA_DE_MORAES","CAROLINA_ANDRADE","MATHEUS_LOPES","SAMUEL_PINTO"]
pdi_js = ["const PDI_DATA = {"]
for sheet in sheets_pdi:
    if sheet not in wb3.sheetnames: continue
    ws3 = wb3[sheet]
    rows3 = [r for r in ws3.iter_rows(values_only=True) if any(v is not None for v in r)]
    info = {"colaborador":"","cargo":"","gestor":"","area":"","ciclo":"","periodo":"","itens":[]}
    for i, row in enumerate(rows3):
        if i == 1: info["colaborador"],info["cargo"],info["gestor"] = str(row[1] or ""),str(row[2] or ""),str(row[3] or "")
        elif i == 3: info["area"],info["ciclo"],info["periodo"] = str(row[1] or ""),str(row[2] or ""),str(row[3] or "")
        elif i >= 5 and row[1] and str(row[1]).strip() != "COMPETENCIA":
            prazo = row[9]
            prazo = prazo.strftime("%d/%m/%Y") if hasattr(prazo,"strftime") else str(prazo or "-")
            info["itens"].append({"competencia":str(row[1] or ""),"categoria":str(row[2] or ""),"situacao":str(row[3] or ""),"objetivo":str(row[4] or ""),"acao":str(row[5] or ""),"tipo":str(row[6] or ""),"responsavel":str(row[7] or ""),"prioridade":str(row[8] or ""),"prazo":prazo,"indicador":str(row[10] or ""),"avaliacao":str(row[11] or ""),"status":str(row[12] or ""),"resultado":str(row[13] or ""),"obs":str(row[14] or "") if row[14] else ""})
    pdi_js += [f"  '{sheet}': {{", f"    colaborador:'{esc(info['colaborador'])}', cargo:'{esc(info['cargo'])}', gestor:'{esc(info['gestor'])}',", f"    area:'{esc(info['area'])}', ciclo:'{esc(info['ciclo'])}', periodo:'{esc(info['periodo'])}',", "    itens:["]
    for it in info["itens"]:
        pdi_js.append(f"      {{competencia:'{esc(it['competencia'])}',categoria:'{esc(it['categoria'])}',situacao:'{esc(it['situacao'])}',objetivo:'{esc(it['objetivo'])}',acao:'{esc(it['acao'])}',tipo:'{esc(it['tipo'])}',responsavel:'{esc(it['responsavel'])}',prioridade:'{esc(it['prioridade'])}',prazo:'{esc(it['prazo'])}',indicador:'{esc(it['indicador'])}',avaliacao:'{esc(it['avaliacao'])}',status:'{esc(it['status'])}',resultado:'{esc(it['resultado'])}',obs:'{esc(it['obs'])}'}},")
    pdi_js += ["    ],", "  }},"]
pdi_js.append("};")
content = re.sub(r"const PDI_DATA = \{.*?\n\};", "\n".join(pdi_js), content, flags=re.DOTALL)
print(f"PDI: {len([s for s in sheets_pdi if s in wb3.sheetnames])} analistas")

# ── ESCALA ────────────────────────────────────────────────────────────────────
wb4 = load_workbook(EXCEL_ESCALA, read_only=True, data_only=True)
ws4 = wb4.active
rows4 = list(ws4.iter_rows(values_only=True))
analistas4 = [rows4[i][0] for i in range(1, len(rows4)) if rows4[i][0]]
dates4 = [v.date() if hasattr(v,"date") else v for v in rows4[0][1:] if v]
escala_js = ["const ESCALA_DATA = {"]
for i, nome in enumerate(analistas4):
    row = rows4[i+1]
    entries = {str(d): str(row[j+1]).strip() for j,d in enumerate(dates4) if j+1 < len(row) and row[j+1]}
    escala_js.append(f"  '{nome}': {json.dumps(entries, ensure_ascii=False)},")
escala_js.append("};")
content = re.sub(r"const ESCALA_DATA = \{.*?\n\};", "\n".join(escala_js), content, flags=re.DOTALL)
print(f"Escala: {len(analistas4)} analistas, {len(dates4)} datas")

# ── SALVA ─────────────────────────────────────────────────────────────────────
with open(HTML, "w", encoding="utf-8") as f:
    f.write(content)
print(f"\nDashboard atualizado com sucesso! [{ts}]")
